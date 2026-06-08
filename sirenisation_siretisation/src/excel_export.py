"""Export Excel uniformisé pour les Phases 1/2/3 et la consolidation."""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

LABELS = {
    "VALIDE_FORT":   "Valide_fort",
    "VALIDE":        "Valide",
    "DOUTEUX":       "Douteux",
    "REJETE":        "Rejeté",
    "SANS_SIRET":    "Sans_SIRET",
    "SIRET_INCONNU": "SIRET_inconnu",
    "SANS_SIREN":    "Sans_SIREN",
    "SIREN_INCONNU": "SIREN_inconnu",
    "SANS_CANDIDAT": "Sans_candidat",
}

COULEURS_HEADER = {
    "VALIDE_FORT":   "1A7341",
    "VALIDE":        "2E8B57",
    "DOUTEUX":       "D4A017",
    "REJETE":        "C0392B",
    "SANS_SIRET":    "595959",
    "SIRET_INCONNU": "595959",
    "SANS_SIREN":    "595959",
    "SIREN_INCONNU": "595959",
    "SANS_CANDIDAT": "595959",
}

COULEURS_LIGNES = {
    "VALIDE_FORT":   "E8F5E9",
    "VALIDE":        "F1F8F4",
    "DOUTEUX":       "FFF8E7",
    "REJETE":        "FDEDEC",
    "SANS_SIRET":    "F2F2F2",
    "SIRET_INCONNU": "F2F2F2",
    "SANS_SIREN":    "F2F2F2",
    "SIREN_INCONNU": "F2F2F2",
    "SANS_CANDIDAT": "ECECEC",
}

COULEURS_SYNTH = {
    "Valide_fort":   "C6EFCE",
    "Valide":        "D9F0E3",
    "Douteux":       "FFEB9C",
    "Rejeté":        "FFC7CE",
    "Sans_SIRET":    "E2E2E2",
    "SIRET_inconnu": "ECECEC",
    "Sans_SIREN":    "E2E2E2",
    "SIREN_inconnu": "ECECEC",
    "Sans_candidat": "ECECEC",
    "TOTAL":         "D9E1F2",
}

COULEURS_SYNTH_COH = {
    "EJ avec tous leurs EG cohérents":     "C6EFCE",  
    "EJ avec au moins un EG incohérent":   "FFC7CE",  
    "EJ validés sans aucun EG validé":     "FFEB9C",  
    "EG validés sans EJ validé":           "FFEB9C",  
    "EG orphelins (sans EJ parent)":       "E2E2E2",  
    "Total couples (EG, EJ) cohérents":    "E8F5E9",
    "Total couples (EG, EJ) incohérents":  "FDEDEC",
    "—": "FFFFFF",
}

def _style_header(ws, couleur: str):
    fill = PatternFill("solid", start_color=couleur, end_color=couleur)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30


def _style_lignes(ws, couleur_alt: str):
    fa = PatternFill("solid", start_color=couleur_alt, end_color=couleur_alt)
    fb = PatternFill("solid", start_color="FFFFFF",    end_color="FFFFFF")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        fill = fa if i % 2 == 0 else fb
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws, max_w: int = 45):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, max_w)


def _ecrire_synthese(writer, compteurs: dict, ordre: list):
    total  = sum(compteurs.values())
    lignes = [
        {"Statut": LABELS[s], "Nb": compteurs.get(s, 0),
         "% du total": round(compteurs.get(s, 0) / total * 100, 1) if total else 0.0}
        for s in ordre
    ]
    lignes.append({"Statut": "TOTAL", "Nb": total, "% du total": 100.0})

    pd.DataFrame(lignes).to_excel(writer, sheet_name="Synthèse", index=False)
    ws = writer.sheets["Synthèse"]
    _style_header(ws, "1F4E79")
    for row in ws.iter_rows(min_row=2):
        label = str(row[0].value)
        fond  = COULEURS_SYNTH.get(label, "FFFFFF")
        for cell in row:
            cell.fill      = PatternFill("solid", start_color=fond, end_color=fond)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if label == "TOTAL":
                cell.font = Font(bold=True, name="Arial", size=10)
    _auto_width(ws, max_w=20)
    ws.freeze_panes = "A2"


def export_phase1_excel(
    df_resultats: pd.DataFrame,
    chemin: Path,
    colonnes_complet: list,
    colonnes_info: list,
    statuts_score: list,
    statuts_info: list,
):
    """Export Phase 1 : feuilles par statut (score + info) + Synthèse."""
    chemin.parent.mkdir(parents=True, exist_ok=True)
    compteurs = {}

    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        for statut in statuts_score:
            sous_df = df_resultats[df_resultats["statut"] == statut]
            compteurs[statut] = len(sous_df)
            cols = [c for c in colonnes_complet if c in sous_df.columns and c != "statut"]
            sous_df[cols].to_excel(writer, sheet_name=LABELS[statut], index=False)
            ws = writer.sheets[LABELS[statut]]
            _style_header(ws, COULEURS_HEADER[statut])
            _style_lignes(ws, COULEURS_LIGNES[statut])
            _auto_width(ws)
            ws.freeze_panes = "A2"

        for statut in statuts_info:
            sous_df = df_resultats[df_resultats["statut"] == statut]
            compteurs[statut] = len(sous_df)
            cols = [c for c in colonnes_info if c in sous_df.columns]
            sous_df[cols].to_excel(writer, sheet_name=LABELS[statut], index=False)
            ws = writer.sheets[LABELS[statut]]
            _style_header(ws, COULEURS_HEADER[statut])
            _style_lignes(ws, COULEURS_LIGNES[statut])
            _auto_width(ws)
            ws.freeze_panes = "A2"

        _ecrire_synthese(writer, compteurs, statuts_score + statuts_info)

    return compteurs


def export_topn_excel(df_topn: pd.DataFrame, chemin: Path, colonnes: list,
                      sheet_name: str = "TopN"):
    """Export Phase 2/3 : feuille TopN + Synthèse (basée sur les rang 1)."""
    chemin.parent.mkdir(parents=True, exist_ok=True)

    cols = [c for c in colonnes if c in df_topn.columns]
    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        df_topn[cols].to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _style_header(ws, "1F4E79")
        _style_lignes(ws, "F2F6FC")
        _auto_width(ws)
        ws.freeze_panes = "A2"

        rang1 = df_topn[df_topn["rang"] == 1]
        compteurs = {
            s: int((rang1["statut_candidat"] == s).sum())
            for s in ("VALIDE_FORT", "VALIDE", "DOUTEUX", "REJETE", "SANS_CANDIDAT")
        }

        total  = sum(compteurs.values())
        lignes = [
            {"Statut": LABELS[s], "Nb": compteurs[s],
             "% du total": round(compteurs[s] / total * 100, 1) if total else 0.0}
            for s in compteurs
        ]
        lignes.append({"Statut": "TOTAL", "Nb": total, "% du total": 100.0})

        pd.DataFrame(lignes).to_excel(writer, sheet_name="Synthèse", index=False)
        ws = writer.sheets["Synthèse"]
        _style_header(ws, "1F4E79")
        for row in ws.iter_rows(min_row=2):
            label = str(row[0].value)
            fond  = COULEURS_SYNTH.get(label, "FFFFFF")
            for cell in row:
                cell.fill      = PatternFill("solid", start_color=fond, end_color=fond)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if label == "TOTAL":
                    cell.font = Font(bold=True, name="Arial", size=10)
        _auto_width(ws, max_w=20)
        ws.freeze_panes = "A2"

    return compteurs


# ─── Export comparaison vue A (par EJ et ses EG) ─────────────────────────────

# Couleurs réutilisées pour les onglets de cohérence
COULEUR_HEADER_COHERENT   = "1A7341"
COULEUR_HEADER_INCOHERENT = "C0392B"
COULEUR_HEADER_PARTIEL    = "D4A017"
COULEUR_HEADER_SYNTHESE   = "1F4E79"

COULEUR_LIGNES_COHERENT   = "E8F5E9"
COULEUR_LIGNES_INCOHERENT = "FDEDEC"
COULEUR_LIGNES_PARTIEL    = "FFF8E7"


# Colonnes ordonnées par bloc logique
COLS_VUE_A = [
    # Identifiants EJ
    "nmfinessej_ej",
    "siren_finess_ej",
    "nom_ej",
    "commune_ej",
    # Résultats sirenisation
    "siren_ej_retenu",
    "statut_ej",
    "phase_ej",
    "nom_ul_retenu",
    "denomination_ul",
    # Identifiants EG
    "nmfinessej_eg",
    "nmfinessetab_stru",
    "nom_eg",
    "commune_eg",
    "siret_finess_eg",
    # Résultats siretisation
    "siret_eg_retenu",
    "siren_du_siret_eg",
    "statut_eg",
    "phase_eg",
    "nom_etab_retenu",
    "adresse_etab_sirene",
    # Verdict
    "statut_coherence",
]


def export_comparaison_excel(
    df_vue_a: pd.DataFrame,
    chemin: Path,
    colonnes: list = None,
):
    """
    Export de la vue A en 4 feuilles : Synthese, Coherent, Incoherent, Partiel.

    Pour les feuilles Coherent et Incoherent : couples (EG, EJ) où les deux
    sont validés → toutes les colonnes sont remplies.

    Pour la feuille Partiel : regroupe PARTIEL_EG (EG validé sans EJ validé)
    et PARTIEL_EJ (EJ validé sans EG validé). Une colonne 'cote_valide'
    précise lequel des deux est validé.
    """
    chemin.parent.mkdir(parents=True, exist_ok=True)

    if colonnes is None:
        colonnes = COLS_VUE_A

    df = df_vue_a.copy()
    cols = [c for c in colonnes if c in df.columns]

    df_coherent   = df[df["statut_coherence"] == "COHERENT"][cols]
    df_incoherent = df[df["statut_coherence"] == "INCOHERENT"][cols]

    # Partiel : on regroupe PARTIEL_EG + PARTIEL_EJ + ORPHELIN avec colonne cote_valide
    df_partiel = df[df["statut_coherence"].isin(["PARTIEL_EG", "PARTIEL_EJ", "ORPHELIN"])].copy()
    df_partiel["cote_valide"] = df_partiel["statut_coherence"].map({
        "PARTIEL_EG": "EG seul validé",
        "PARTIEL_EJ": "EJ seul validé",
        "ORPHELIN":   "EG sans EJ parent",
    })
    cols_partiel = ["cote_valide"] + [c for c in cols if c != "statut_coherence"]
    cols_partiel = [c for c in cols_partiel if c in df_partiel.columns]
    df_partiel = df_partiel[cols_partiel]

    # ─── Compteurs au niveau EJ ──────────────────────────────────────────────
    df_avec_match = df[df["statut_coherence"].isin(["COHERENT", "INCOHERENT"])]
    par_ej = df_avec_match.groupby("nmfinessej_eg")["statut_coherence"].apply(set) \
        if "nmfinessej_eg" in df_avec_match.columns else pd.Series(dtype=object)

    nb_ej_tous_coh    = sum(1 for s in par_ej if s == {"COHERENT"})
    nb_ej_avec_incoh  = sum(1 for s in par_ej if "INCOHERENT" in s)
    nb_partiel_ej     = int((df["statut_coherence"] == "PARTIEL_EJ").sum())
    nb_partiel_eg     = int((df["statut_coherence"] == "PARTIEL_EG").sum())
    nb_orphelins      = int((df["statut_coherence"] == "ORPHELIN").sum())
    nb_couples_coh    = int((df["statut_coherence"] == "COHERENT").sum())
    nb_couples_incoh  = int((df["statut_coherence"] == "INCOHERENT").sum())

    lignes_synth = [
        {"Indicateur": "EJ avec tous leurs EG cohérents",     "Nombre": nb_ej_tous_coh},
        {"Indicateur": "EJ avec au moins un EG incohérent",   "Nombre": nb_ej_avec_incoh},
        {"Indicateur": "EJ validés sans aucun EG validé",     "Nombre": nb_partiel_ej},
        {"Indicateur": "EG validés sans EJ validé",           "Nombre": nb_partiel_eg},
        {"Indicateur": "EG orphelins (sans EJ parent)",       "Nombre": nb_orphelins},
        {"Indicateur": "—",                                   "Nombre": ""},
        {"Indicateur": "Total couples (EG, EJ) cohérents",    "Nombre": nb_couples_coh},
        {"Indicateur": "Total couples (EG, EJ) incohérents",  "Nombre": nb_couples_incoh},
    ]
    df_synth = pd.DataFrame(lignes_synth)

    # ─── Écriture Excel ──────────────────────────────────────────────────────
    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:

        # Feuille 1 : Synthese
        df_synth.to_excel(writer, sheet_name="Synthese", index=False)
        ws = writer.sheets["Synthese"]
        _style_header(ws, COULEUR_HEADER_SYNTHESE)
        for row in ws.iter_rows(min_row=2):
            label = str(row[0].value)
        
            fond = COULEURS_SYNTH_COH.get(label, "FFFFFF")
        
            for cell in row:
                cell.fill = PatternFill("solid", start_color=fond, end_color=fond)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
                # Mise en valeur des totaux
                if "Total" in label:
                    cell.font = Font(bold=True)
        _auto_width(ws, max_w=45)
        ws.freeze_panes = "A2"

        # Feuille 2 : Coherent
        if len(df_coherent) > 0:
            df_coherent.to_excel(writer, sheet_name="Coherent", index=False)
            ws = writer.sheets["Coherent"]
            _style_header(ws, COULEUR_HEADER_COHERENT)
            _style_lignes(ws, COULEUR_LIGNES_COHERENT)
            _auto_width(ws)
            ws.freeze_panes = "A2"

        # Feuille 3 : Incoherent
        if len(df_incoherent) > 0:
            df_incoherent.to_excel(writer, sheet_name="Incoherent", index=False)
            ws = writer.sheets["Incoherent"]
            _style_header(ws, COULEUR_HEADER_INCOHERENT)
            _style_lignes(ws, COULEUR_LIGNES_INCOHERENT)
            _auto_width(ws)
            ws.freeze_panes = "A2"

        # Feuille 4 : Partiel
        if len(df_partiel) > 0:
            df_partiel.to_excel(writer, sheet_name="Partiel", index=False)
            ws = writer.sheets["Partiel"]
            _style_header(ws, COULEUR_HEADER_PARTIEL)
            _style_lignes(ws, COULEUR_LIGNES_PARTIEL)
            _auto_width(ws)
            ws.freeze_panes = "A2"

    return {
        "EJ_tous_coherents":   nb_ej_tous_coh,
        "EJ_avec_incoherent":  nb_ej_avec_incoh,
        "PARTIEL_EJ":          nb_partiel_ej,
        "PARTIEL_EG":          nb_partiel_eg,
        "ORPHELIN":            nb_orphelins,
        "COHERENT_couples":    nb_couples_coh,
        "INCOHERENT_couples":  nb_couples_incoh,
    }


# ─── Anciens dicts conservés pour rétrocompatibilité ─────────────────────────

LABELS_COHERENCE = {
    "COHERENT":   "Cohérent",
    "INCOHERENT": "Incohérent",
    "PARTIEL_EG": "Partiel_EG",
    "PARTIEL_EJ": "Partiel_EJ",
    "ORPHELIN":   "Orphelin",
}
